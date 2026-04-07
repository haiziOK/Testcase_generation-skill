#!/usr/bin/env python
"""Verify generated Excel file."""
import openpyxl
import sys
import io

# Fix encoding issues on Windows with GBK console
if hasattr(sys.stdout, 'buffer') and (sys.stdout.encoding is None or sys.stdout.encoding.upper() not in ('UTF-8', 'UTF8')):
    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    except Exception:
        pass  # Keep original stdout if any error

def verify_excel(file_path):
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        print(f"Sheet name: {ws.title}")
        print(f"Total rows: {ws.max_row}")
        print(f"Total columns: {ws.max_column}")

        # Print headers
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column+1)]
        print("Headers:", headers)

        # Print first few test cases
        print("\nFirst 3 test cases:")
        for r in range(2, min(5, ws.max_row+1)):
            row_data = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column+1)]
            print(f"Row {r}: {row_data}")

    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)

if __name__ == '__main__':
    verify_excel('examples/test_cases.xlsx')