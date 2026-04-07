"""
Excel writer for test cases.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any
import logging

logger = logging.getLogger(__name__)

class ExcelWriter:
    """Write test cases to Excel file."""

    def __init__(self):
        # Define styles
        self.header_font = Font(bold=True, size=12, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center")

        self.cell_font = Font(size=11)
        self.cell_alignment = Alignment(vertical="top", wrap_text=True)

        self.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

        self.priority_colors = {
            "high": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
            "medium": PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid"),
            "low": PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
        }

    def write_test_cases(self, test_cases: List[Dict], output_path: str):
        """Write test cases to Excel file."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "测试用例"

        # Define column headers
        headers = [
            "测试用例ID",
            "标题",
            "描述",
            "步骤",
            "预期结果",
            "优先级",
            "需求ID",
            "状态",
            "备注"
        ]

        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border

        # Write test cases
        for row_idx, test_case in enumerate(test_cases, start=2):
            # Test Case ID
            ws.cell(row=row_idx, column=1, value=test_case.get("id", f"TC{row_idx-1:03d}"))

            # Title
            ws.cell(row=row_idx, column=2, value=test_case.get("title", ""))

            # Description
            ws.cell(row=row_idx, column=3, value=test_case.get("description", ""))

            # Steps (join list with newlines)
            steps = test_case.get("steps", [])
            steps_text = "\n".join([f"{i+1}. {step}" for i, step in enumerate(steps)])
            ws.cell(row=row_idx, column=4, value=steps_text)

            # Expected Result
            ws.cell(row=row_idx, column=5, value=test_case.get("expected_result", ""))

            # Priority
            priority = test_case.get("priority", "medium").lower()
            priority_cell = ws.cell(row=row_idx, column=6, value=priority.capitalize())
            if priority in self.priority_colors:
                priority_cell.fill = self.priority_colors[priority]

            # Requirement ID
            ws.cell(row=row_idx, column=7, value=test_case.get("requirement_id", ""))

            # Status (default: Not Executed)
            ws.cell(row=row_idx, column=8, value="Not Executed")

            # Notes (empty)
            ws.cell(row=row_idx, column=9, value="")

            # Apply formatting to all cells in this row
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = self.cell_font
                cell.alignment = self.cell_alignment
                cell.border = self.border

        # Auto-adjust column widths
        self._auto_adjust_columns(ws)

        # Save workbook
        wb.save(output_path)
        logger.info(f"Test cases written to {output_path}")

    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths based on content."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            # Find the max length in the column
            for cell in column:
                try:
                    if cell.value:
                        # Count length, considering newlines
                        lines = str(cell.value).split('\n')
                        max_line_length = max(len(line) for line in lines)
                        if max_line_length > max_length:
                            max_length = max_line_length
                except:
                    pass

            # Set column width with some padding
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width

    def write_from_document(self, document_path: str, output_path: str, max_cases_per_req: int = 5):
        """Generate test cases from document and write to Excel."""
        from testcase_generator import TestCaseGenerator

        generator = TestCaseGenerator()
        test_cases = generator.generate_from_document(document_path, max_cases_per_req)
        self.write_test_cases(test_cases, output_path)